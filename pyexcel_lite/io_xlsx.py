"""XLSX and CSV persistence."""

from __future__ import annotations

import csv
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import column_index_from_string

from .cell_address import index_to_column_name
from .workbook import CellData, CellStyle, WorkbookData, WorksheetData


def save_xlsx(workbook_data: WorkbookData, path: str | Path) -> None:
    target = Path(path)
    book = Workbook()
    default = book.active
    book.remove(default)
    for sheet_data in workbook_data.sheets:
        sheet = book.create_sheet(sheet_data.name)
        for row, column, cell_data in sheet_data.iter_used_cells():
            excel_cell = sheet.cell(row=row + 1, column=column + 1, value=cell_data.value)
            apply_style_to_openpyxl(excel_cell, cell_data.style)
        for column, width in sheet_data.column_widths.items():
            sheet.column_dimensions[index_to_column_name(column)].width = width
        for row, height in sheet_data.row_heights.items():
            sheet.row_dimensions[row + 1].height = height
    book.active = workbook_data.active_sheet_index
    book.save(target)
    workbook_data.path = str(target)


def load_xlsx(path: str | Path) -> WorkbookData:
    source = Path(path)
    book = load_workbook(source, data_only=False)
    sheets: list[WorksheetData] = []
    for openpyxl_sheet in book.worksheets:
        sheet_data = WorksheetData(
            name=openpyxl_sheet.title,
            row_count=max(200, openpyxl_sheet.max_row or 1),
            column_count=max(52, openpyxl_sheet.max_column or 1),
        )
        for row in openpyxl_sheet.iter_rows():
            for excel_cell in row:
                if excel_cell.value is None and not excel_cell.has_style:
                    continue
                style = style_from_openpyxl(excel_cell)
                row_index = excel_cell.row - 1
                column_index = excel_cell.column - 1
                value = "" if excel_cell.value is None else excel_cell.value
                sheet_data.cells[(row_index, column_index)] = CellData(value=value, style=style)
                sheet_data.track_formula_cell(row_index, column_index, value)
        for key, dimension in openpyxl_sheet.column_dimensions.items():
            if dimension.width:
                sheet_data.column_widths[column_index_from_string(key) - 1] = int(dimension.width)
        for row_index, dimension in openpyxl_sheet.row_dimensions.items():
            if dimension.height:
                sheet_data.row_heights[row_index - 1] = int(dimension.height)
        sheets.append(sheet_data)
    workbook_data = WorkbookData(sheets=sheets or [WorksheetData(name="Sheet1")], active_sheet_index=book.index(book.active))
    workbook_data.path = str(source)
    return workbook_data


def export_csv(sheet_data: WorksheetData, path: str | Path) -> None:
    max_row = 0
    max_column = 0
    for row, column, cell in sheet_data.iter_used_cells():
        if not cell.is_empty:
            max_row = max(max_row, row)
            max_column = max(max_column, column)
    with Path(path).open("w", newline="", encoding="utf-8-sig") as handle:
        writer = csv.writer(handle)
        for row in range(max_row + 1):
            writer.writerow([sheet_data.raw_value(row, column) for column in range(max_column + 1)])


def apply_style_to_openpyxl(excel_cell: Any, style: CellStyle) -> None:
    excel_cell.font = Font(
        name=style.font_family,
        size=style.font_size,
        bold=style.bold,
        italic=style.italic,
        underline="single" if style.underline else None,
        color=style.text_color.replace("#", ""),
    )
    if style.fill_color.lower() != "#ffffff":
        excel_cell.fill = PatternFill("solid", fgColor=style.fill_color.replace("#", ""))
    excel_cell.alignment = Alignment(
        horizontal=None if style.horizontal == "general" else style.horizontal,
        vertical=style.vertical,
        wrap_text=style.wrap_text,
    )
    if style.number_format != "General":
        excel_cell.number_format = style.number_format


def style_from_openpyxl(excel_cell: Any) -> CellStyle:
    font = excel_cell.font
    fill = excel_cell.fill
    alignment = excel_cell.alignment
    text_color = "#111827"
    if font and font.color and font.color.type == "rgb" and font.color.rgb:
        text_color = "#" + font.color.rgb[-6:]
    fill_color = "#ffffff"
    if fill and fill.fill_type == "solid" and fill.fgColor and fill.fgColor.type == "rgb" and fill.fgColor.rgb:
        fill_color = "#" + fill.fgColor.rgb[-6:]
    return CellStyle(
        font_family=font.name or "Segoe UI",
        font_size=int(font.sz or 10),
        bold=bool(font.bold),
        italic=bool(font.italic),
        underline=bool(font.underline),
        text_color=text_color,
        fill_color=fill_color,
        horizontal=alignment.horizontal or "general",
        vertical=alignment.vertical or "center",
        number_format=excel_cell.number_format or "General",
        wrap_text=bool(alignment.wrap_text),
    )
